import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ===== STYLES =====
BLACK = "1A1A1A"
RED = "E31B23"
CREAM = "FFF8F0"
LIGHT = "F5EDE0"
WHITE = "FFFFFF"
GRAY = "EFEFEF"
DARK_GRAY = "555555"

thin = Side(border_style="thin", color="DADADA")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def style_header(cell, fill=RED, font_color=WHITE, size=12):
    cell.fill = PatternFill("solid", fgColor=fill)
    cell.font = Font(bold=True, color=font_color, size=size, name="Calibri")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

def style_cell(cell, fill=WHITE, bold=False, size=10, color=BLACK, wrap=True, align="left"):
    cell.fill = PatternFill("solid", fgColor=fill)
    cell.font = Font(bold=bold, color=color, size=size, name="Calibri")
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    cell.border = border

def title_row(ws, row, text, span, fill=BLACK, color=WHITE, size=14):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1, value=text)
    c.fill = PatternFill("solid", fgColor=fill)
    c.font = Font(bold=True, color=color, size=size, name="Calibri")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 28

# ===== SHEET 1: COVER =====
ws = wb.active
ws.title = "Overview"
ws.sheet_view.showGridLines = False

ws.column_dimensions['A'].width = 4
ws.column_dimensions['B'].width = 28
ws.column_dimensions['C'].width = 70
ws.column_dimensions['D'].width = 4

ws.merge_cells("B2:C2")
c = ws.cell(row=2, column=2, value="WOK!N — INSTAGRAM CONTENT CALENDAR")
c.fill = PatternFill("solid", fgColor=BLACK)
c.font = Font(bold=True, color=WHITE, size=20, name="Calibri")
c.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[2].height = 44

ws.merge_cells("B3:C3")
c = ws.cell(row=3, column=2, value="Bold Flavours. Zero Compromises. Built to go viral.")
c.fill = PatternFill("solid", fgColor=RED)
c.font = Font(bold=True, color=WHITE, size=11, name="Calibri", italic=True)
c.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[3].height = 24

overview_data = [
    ("BRAND VOICE", "Bold, slightly chaotic, fast-paced, self-aware. Wendy's Twitter meets a Bangkok street kitchen."),
    ("PLATFORM", "Instagram (Reels + Stories primary, Carousels secondary)"),
    ("POSTING CADENCE", "1 Reel/day + 3-5 Stories/day + 2 Carousels/week"),
    ("BEST TIMES (PKT)", "Reels: 1PM, 8PM, 11PM | Stories: 12PM, 6PM, 10PM"),
    ("HASHTAG RULE", "8-12 max per post. Never 30. Mix brand + niche + trending."),
    ("KEY METRIC", "Saves + Shares > Likes. Save rate target: 5%+"),
    ("", ""),
    ("CONTENT PILLARS", ""),
    ("1. WOK CAM", "POV, macro shots, fire, oil, sizzle. ASMR. Visual ASMR."),
    ("2. CHEF UNHINGED", "Chef commentary, hot takes, kitchen drama, ranking dishes."),
    ("3. CUSTOMER CHAOS", "Real reactions, dares, challenges, first-bite content."),
    ("4. STREET SMART", "Pan Asian food facts, myth busting, did-you-knows."),
    ("5. MEME LORD", "Trending audio, restaurant memes, food humour."),
    ("", ""),
    ("SHEETS IN THIS WORKBOOK", ""),
    ("Daily Reels", "30-day reel calendar with hooks, captions, and shoot notes"),
    ("Daily Stories", "Daily story plan — 5 stories per day for 30 days"),
    ("Backup Reels", "50+ extra reel ideas if you can't shoot the planned one"),
    ("Backup Stories", "60+ extra story ideas — quick to shoot, low effort"),
    ("Hooks Library", "Stop-the-scroll hook formulas + caption templates"),
    ("Hashtag Bank", "Pre-built hashtag sets by category"),
    ("Tracker", "Performance tracking template"),
]

for i, (k, v) in enumerate(overview_data, start=5):
    if k == "" and v == "":
        ws.row_dimensions[i].height = 8
        continue
    if v == "" and k != "":
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=3)
        c = ws.cell(row=i, column=2, value=k)
        c.fill = PatternFill("solid", fgColor=LIGHT)
        c.font = Font(bold=True, color=BLACK, size=11)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[i].height = 22
    else:
        c1 = ws.cell(row=i, column=2, value=k)
        c1.fill = PatternFill("solid", fgColor=BLACK)
        c1.font = Font(bold=True, color=WHITE, size=10)
        c1.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
        c2 = ws.cell(row=i, column=3, value=v)
        c2.fill = PatternFill("solid", fgColor=WHITE)
        c2.font = Font(color=BLACK, size=10)
        c2.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
        c1.border = border
        c2.border = border
        ws.row_dimensions[i].height = 24

print("Overview sheet done")

# ===== SHEET 2: DAILY REELS =====
ws = wb.create_sheet("Daily Reels")
ws.sheet_view.showGridLines = False

headers = ["Day", "Date", "Weekday", "Pillar", "Hook (First 1.5s)", "Reel Concept", "Caption", "Sound/Audio", "Shoot Notes", "Status"]
widths = [6, 12, 11, 16, 32, 50, 45, 22, 30, 12]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

title_row(ws, 1, "30-DAY REEL CALENDAR — Wok!n", len(headers), fill=BLACK)
title_row(ws, 2, "One reel per day. Plan your shoots in batches of 5-7 days.", len(headers), fill=RED, size=10)

for i, h in enumerate(headers, 1):
    style_header(ws.cell(row=3, column=i, value=h))
ws.row_dimensions[3].height = 32

reels = [
    (1, "MON", "Wok Cam", "POV: You ordered the chilli garlic chicken.", "Macro slow-mo of wok flames + chicken hitting oil. 0.5x speed first, then real time.", "Your eardrums called. They want this on repeat. 🔊🔥", "Trending bass drop", "Tripod, side angle, kitchen torch on standby"),
    (2, "TUE", "Chef Unhinged", "Chef ranks our menu from 'mild lie' to 'will end you'.", "Chef pulls dishes one by one, deadpan face, rates spice. End on the killer.", "Chef said 'rank them or fire me.' We let him cook. 🌶️", "Trending POV audio", "Clean counter, 5-7 dishes pre-plated"),
    (3, "WED", "Street Smart", "3 Pan Asian myths Pakistanis still believe.", "Fast cuts, text overlay, chef debunks with eye roll. 'Chow mein is Chinese' WRONG.", "We said what we said. 🥢", "Quick beat, no vocals", "Use teleprompter or notes off-cam"),
    (4, "THU", "Customer Chaos", "We let a 6-year-old design our 'kid menu' for a day.", "Cute kid points at dishes, chef reacts. Subtitled. Wholesome chaos.", "Toddler-approved. Adult-tested. 🧒🍜", "Cute upbeat indie", "Need parental release form, shoot in daylight hours"),
    (5, "FRI", "Meme Lord", "Things I order that say more about me than my therapist.", "Carousel-as-reel: dishes with personality labels. Trending audio.", "Your order = your personality. We don't make the rules. 💅", "Latest viral audio", "Shoot all 6-8 dish closeups in one session"),
    (6, "SAT", "Wok Cam", "60 seconds. 1 wok. 12 dishes.", "Speed-run montage. End frame: 'Now imagine eating all of this.'", "We don't slow down. Neither should you. ⚡", "Fast electronic", "Shoot during prep, batch all wok shots"),
    (7, "SUN", "Chef Unhinged", "Sunday brunch but make it Pan Asian.", "Chef plates dim sum + paratha + chilli oil combo nobody asked for.", "Brunch rules are made to be wok'd. 🍳🥟", "Chill lo-fi", "Bright morning light, white plates"),
    (8, "MON", "Street Smart", "What 'Szechuan' actually means (and why your mouth tingles).", "Animated text + chef demo with peppercorns. End: 'Now you know.'", "Spice education. Free of charge. 🌶️📚", "Documentary-style", "Closeup of peppercorns, ring light"),
    (9, "TUE", "Customer Chaos", "We gave strangers our hottest dish. Watch their souls leave.", "Reactions montage. Make it dramatic. 5-6 different people.", "RIP to all 7 of them. Worth it. 💀🔥", "Dramatic build-up", "Need 6+ willing volunteers, water on standby"),
    (10, "WED", "Meme Lord", "Wokin menu items as red flags.", "Carousel-as-reel: 10 dishes matched to dating personality types.", "We're not therapists but we feed you anyway. 🚩", "Trending sassy audio", "Studio plate shots, red overlay graphics"),
    (11, "THU", "Wok Cam", "ASMR: chef chops garlic for 47 seconds.", "No music. Just sound. Macro lens. Wood board.", "Turn the volume up. Thank us later. 🎧🧄", "RAW SOUND ONLY", "Quiet kitchen, lavalier mic close to board"),
    (12, "FRI", "Chef Unhinged", "Chef guesses what dish a customer ordered just from how they walked in.", "Comedy bit. Chef studies customer, makes call, dish appears.", "Chef has powers. We're a little scared. 👁️", "Sherlock-style audio", "Need 3-4 actors/regulars, front-of-house shoot"),
    (13, "SAT", "Customer Chaos", "Customers describe our food in ONE word.", "Fast cuts. Last word is hilarious like 'religious' or 'illegal.'", "Word on the street. 🗣️", "Hip beat", "Stop diners after they finish, golden hour"),
    (14, "SUN", "Street Smart", "Why we use THIS exact wok (and what cheap ones get wrong).", "Show seasoning, wok hei, the science. Vox-explainer style.", "Wok talk. Nerd alert. 🥢🔬", "Inquisitive instrumental", "Detail shots of wok, B-roll of cheap pan comparison"),
    (15, "MON", "Meme Lord", "Tell me you're a Wokin regular without telling me.", "Trending format. Use real customers if possible.", "We see you. Every. Single. Time. 👀", "Trending audio", "Recruit 4-5 regulars, quick shoot"),
    (16, "TUE", "Wok Cam", "The exact moment fire meets oil.", "Slow-mo, 240fps if possible. ONE shot, ONE moment. Cinematic.", "Goosebumps. Every. Time. 🔥", "Drop sound only", "High-speed camera or phone slow-mo, fire safety!"),
    (17, "WED", "Challenge Launch ⭐", "INTRODUCING: The Wokin Heat Test.", "Show the dish menacingly. Rules: finish in 10 mins, get it FREE + wall of fame.", "Most won't make it past minute 4. Will you? 📛", "Epic trailer audio", "Build a 'wall of fame' physical board, dramatic lighting"),
    (18, "THU", "Customer Chaos", "First Heat Test attempt.", "Real customer crying, sweating, laughing. Subtitled chaos.", "We have a winner. We also have casualties. ⚰️🌶️", "Tense build", "Multi-cam: close-up + wide, get release form"),
    (19, "FRI", "Street Smart", "Why noodles are slurped (and why your mom was wrong).", "Cultural fact, end with chef demoing the LOUDEST slurp possible.", "Permission granted. SLURP LOUD. 🍜📢", "Funny zoom-in audio", "Lavalier mic close to mouth (yes really)"),
    (20, "SAT", "Meme Lord", "TREND HIJACK — Wok edition.", "Whatever audio is #1 that week, force-fit a Wokin angle. Post within 24h.", "Match the trend tone", "WHATEVER IS TRENDING", "Keep equipment ready, react fast"),
    (21, "SUN", "Chef Unhinged", "Chef reacts to viral 'Asian food' TikToks.", "Greenscreen format. Chef annoyed, then approves one. Pure entertainment.", "Chef has notes. Many notes. 📋😤", "Original audio", "Greenscreen setup or duet format"),
    (22, "MON", "Wok Cam", "Tour of one full ticket — order to plate in 90 seconds.", "POV, hands only, fast paced. Restaurant orchestra vibe.", "This happens 200 times a day. Madness. 🎟️🍳", "Building tempo track", "Shoot during peak service, GoPro on chest"),
    (23, "TUE", "Customer Chaos", "We surprised this regular with their order before they ordered.", "Hidden cam style. Reaction = gold.", "We know what you want. It's a little scary. 🔮", "Mysterious audio", "Coordinate with regular's friend, hide camera"),
    (24, "WED", "Meme Lord", "Wokin dishes as your 2026 horoscope.", "Carousel: 12 slides, one per zodiac sign. Built for screenshots.", "Tag your sign. We don't take responsibility. ♈🍤", "Mystic audio", "Studio plate shots + zodiac graphics"),
    (25, "THU", "Street Smart", "Pakistani vs Pan Asian spice — we tested it.", "Scoville chart, real demo, surprising results.", "Plot twist nobody saw coming. 📊🌶️", "Documentary build", "Need scoville chart graphic + 4 chillies"),
    (26, "FRI", "Chef Unhinged", "Chef makes the menu's most underrated dish.", "Genuine, slightly offended energy. 'You sleep on this. Stop.'", "Order it next time. We dare you. 😤🥢", "Slow build emotional", "Long takes, let chef talk"),
    (27, "SAT", "Wok Cam + UGC", "Best customer-tagged videos this month.", "Compilation reel. Music, energy, chaos.", "You ate. You filmed. We immortalized. 🎬❤️", "Hype anthem", "Save all UGC throughout month, edit Friday"),
    (28, "SUN", "Customer Chaos", "Blind taste test: 5 customers blindfolded.", "Pure entertainment. Pure chaos.", "Number 3 had us SCREAMING. 🙈", "Comedy build-up", "Need 5 volunteers, blindfolds, 5 mystery dishes"),
    (29, "MON", "Meme Lord", "Things only Wokin regulars know.", "Inside-joke energy. Builds cult feel.", "If you know, you know. 🤝", "Trending audio", "Quick montage, 8-10 'inside' references"),
    (30, "TUE", "Community Payoff", "30 days. Here's what YOU made happen.", "Stats reel — followers, dishes served, Heat Test winners. Music. Gratitude.", "From our wok to your table — and back again. ❤️🔥", "Emotional anthem", "Pull all stats from IG insights, design slides"),
]

start_row = 4
import datetime
start_date = datetime.date(2026, 4, 13)
for idx, (day, wkd, pillar, hook, concept, caption, audio, notes) in enumerate(reels):
    row = start_row + idx
    date = start_date + datetime.timedelta(days=idx)
    fill = CREAM if idx % 2 == 0 else WHITE
    pillar_colors = {
        "Wok Cam": "FFE5E5", "Chef Unhinged": "FFF0D9", "Customer Chaos": "E5F4FF",
        "Street Smart": "E8F5E9", "Meme Lord": "F3E5F5", "Challenge Launch ⭐": "FFD6D6",
        "Wok Cam + UGC": "FFE5E5", "Community Payoff": "FFF0D9"
    }
    pf = pillar_colors.get(pillar, fill)
    values = [day, date.strftime("%b %d"), wkd, pillar, hook, concept, caption, audio, notes, ""]
    for i, v in enumerate(values, 1):
        c = ws.cell(row=row, column=i, value=v)
        if i == 4:
            style_cell(c, fill=pf, bold=True, size=9, align="center")
        elif i in (1, 2, 3):
            style_cell(c, fill=fill, bold=True, size=10, align="center")
        elif i == 10:
            style_cell(c, fill="FFF8DC", size=9, align="center")
        else:
            style_cell(c, fill=fill, size=9)
    ws.row_dimensions[row].height = 70

ws.freeze_panes = "A4"
ws.freeze_panes = "A4"
print("Daily Reels sheet done")

# ===== SHEET 3: DAILY STORIES =====
ws = wb.create_sheet("Daily Stories")
ws.sheet_view.showGridLines = False

headers = ["Day", "Date", "Story 1 (Morning)", "Story 2 (Lunch)", "Story 3 (Afternoon)", "Story 4 (Evening)", "Story 5 (Night)"]
widths = [6, 12, 32, 32, 32, 32, 32]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

title_row(ws, 1, "30-DAY STORY PLAN — Wok!n", len(headers), fill=BLACK)
title_row(ws, 2, "5 stories per day. Mix BTS, polls, UGC, countdowns, sneak peeks.", len(headers), fill=RED, size=10)

for i, h in enumerate(headers, 1):
    style_header(ws.cell(row=3, column=i, value=h))
ws.row_dimensions[3].height = 30

stories = [
    ("BTS: Chef opening up the kitchen ☀️", "Poll: Spicy or saucy today?", "Behind the pass: noodles being tossed", "UGC reshare: tag from yesterday", "Tomorrow's reel teaser (2-sec blur)"),
    ("Question sticker: What should we cook tomorrow?", "Quick clip: wok flames close-up", "This or That: Chow mein vs Hakka noodles", "Customer reaction (raw, no edit)", "'Closing soon' countdown"),
    ("Spice Meter Monday: rate your week 🌶️", "Lunch rush time-lapse", "Poll: Tofu or chicken?", "Chef's hand burnt, laughing about it", "End of day total dishes served stat"),
    ("Wok Wednesday teaser graphic", "Mid-prep dim sum folding ASMR", "Quiz: Which dish matches your mood?", "UGC reshare with overlay text", "Late-night menu sneak peek"),
    ("Throwback Thali: old menu item", "Today's secret special reveal", "Vote: Should this stay on menu?", "Customer first-bite reaction", "Tomorrow's drop countdown"),
    ("Friday Flavour Forecast: what's selling out", "Weekend hours reminder graphic", "Behind-the-scenes plating", "Tag-the-friend story", "'Door's open' clip with sizzle sound"),
    ("Saturday energy: kitchen vibes clip", "Question sticker: Order or dine-in?", "Mid-service chaos clip", "UGC compilation slide", "Late dinner empty plates clip"),
    ("Sunday Sauce Talk: chef opens up about a dish", "Brunch combo plating clip", "Poll: Brunch or lunch?", "Customer family wholesome clip", "End of week wrap stat"),
    ("Wokin Wrapped: top dish this week 📊", "New week menu reminder", "BTS: deliveries arriving fresh", "Quick chef hot take", "Countdown to Heat Test (if active)"),
    ("Tuesday teaser: new reel coming", "Quick wok toss video", "Quiz: spice tolerance test", "UGC reshare with sticker", "Bite-sized chef quote"),
    ("Wok Wednesday hype: wok shot", "Lunch ASMR: garlic chopping", "Poll: Which sauce wins?", "Customer order surprise reveal", "Behind-the-pass montage"),
    ("Quick myth-bust slide", "Today's chef pick reveal", "Vote: Mild or volcanic?", "Customer testimonial reshare", "'Tomorrow we go viral' tease"),
    ("Friday FlameFriday graphic", "Pre-rush kitchen calm", "Tag a friend who slurps loud", "Service starting clip", "Weekend special countdown"),
    ("Saturday: doors-open clip", "Behind the chef's pass", "Quiz: Guess the dish from emoji", "Wall of fame Heat Test photo", "Dinner rush time-lapse"),
    ("Sunday Sauce Talk Part 2", "Plating in slow motion", "Poll: Comfort food or adventure?", "Family customer wholesome clip", "Week-recap stats slide"),
    ("Monday Spice Meter rate", "Quick wok flame clip", "This or That dish battle", "UGC reshare", "Reel teaser blur"),
    ("Wokin Trivia: did you know?", "Lunch rush BTS", "Poll: Heat Test winner this week?", "Heat Test prep clip", "Countdown to Wednesday's challenge launch"),
    ("HEAT TEST DAY graphic 🔥", "Live update: who signed up", "Vote: Will they finish it?", "First attempt clip raw", "Winner announcement"),
    ("Heat Test recap slides", "Slurping ASMR clip", "Tag-bait: Who'd survive Heat Test?", "Customer running for water clip", "Tomorrow's reel tease"),
    ("FlameFriday graphic", "Wok flames slow-mo", "Poll: Trend hijack vs original?", "Trend reel reaction tease", "Door's-open Friday energy"),
    ("Saturday brunch tease", "Plating ASMR", "Quiz: Which Wokin dish are you?", "UGC slideshow", "Late-night order clip"),
    ("Sunday Reaction Time graphic", "Greenscreen tease frame", "Vote: Best chef rant of the week?", "Greenscreen behind-the-scenes", "Wrap stats"),
    ("Service start clip", "Order-to-plate timer animation", "Poll: Fastest dish guess?", "Hidden cam tease", "Reel drop reminder"),
    ("Tuesday Tease: hidden cam reveal", "BTS regulars walking in", "This or That regulars edition", "Customer surprised reaction", "Surprise reveal full reel link"),
    ("Horoscope teaser: 'find yours tomorrow'", "Plate spinning clip", "Poll: Believe in horoscopes?", "Carousel sneak peek", "Drop countdown"),
    ("Spice test setup clip", "Scoville chart reveal slide", "Quiz: Which is hotter?", "Test results tease", "Reel link in story"),
    ("Underrated dish reveal graphic", "Plating closeup", "Vote: Tried this dish before?", "Chef passionate clip", "Order reminder sticker"),
    ("UGC Saturday: shoutouts begin", "Customer compilation slide 1", "Tag yourself sticker", "Customer compilation slide 2", "Reel link drop"),
    ("Blind taste test setup", "Volunteers prep clip", "Poll: Will they guess right?", "First reaction clip", "Full reel link"),
    ("Insider day: 'things only regulars know'", "Inside joke graphic 1", "Poll: Are YOU a regular?", "Inside joke graphic 2", "Reel link drop"),
]

start_date_s = datetime.date(2026, 4, 13)
for idx, sts in enumerate(stories):
    row = 4 + idx
    date = start_date_s + datetime.timedelta(days=idx)
    fill = CREAM if idx % 2 == 0 else WHITE
    style_cell(ws.cell(row=row, column=1, value=idx+1), fill=fill, bold=True, align="center")
    style_cell(ws.cell(row=row, column=2, value=date.strftime("%b %d")), fill=fill, bold=True, align="center", size=9)
    for j, s in enumerate(sts):
        style_cell(ws.cell(row=row, column=3+j, value=s), fill=fill, size=9)
    ws.row_dimensions[row].height = 50

ws.freeze_panes = "C4"
print("Daily Stories sheet done")

# ===== SHEET 4: BACKUP REELS =====
ws = wb.create_sheet("Backup Reels")
ws.sheet_view.showGridLines = False

headers = ["#", "Pillar", "Reel Idea", "Effort (1-5)", "Why It Works"]
widths = [6, 18, 60, 12, 50]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

title_row(ws, 1, "BACKUP REEL IDEAS — Use When You Can't Shoot the Plan", len(headers), fill=BLACK)
title_row(ws, 2, "60+ low-effort to mid-effort reel concepts. Pull from this bank anytime.", len(headers), fill=RED, size=10)

for i, h in enumerate(headers, 1):
    style_header(ws.cell(row=3, column=i, value=h))
ws.row_dimensions[3].height = 28

backup_reels = [
    ("Wok Cam", "Single oil-pour shot, slow-mo, no music — let the sizzle do the work", 1, "Pure ASMR. Saves rate goes nuts."),
    ("Wok Cam", "Macro of garlic + chilli hitting hot oil, frame-by-frame", 1, "Visual hook in 0.5 seconds"),
    ("Wok Cam", "Dish plating from above, hands only, 30 sec", 1, "POV format always wins"),
    ("Wok Cam", "Spice tossing in air, slow motion catch", 2, "Visually impressive flex"),
    ("Wok Cam", "Wok flame ignition mega close-up", 1, "Easy to shoot, dramatic"),
    ("Wok Cam", "Steam rising from a fresh dim sum basket", 1, "Sensory, calming"),
    ("Wok Cam", "Sauce being drizzled in slow-mo on a finished dish", 1, "Money shot, hypnotic"),
    ("Wok Cam", "Cleaver chopping vegetables in rhythm to a beat", 2, "Synced cut = engaging"),
    ("Wok Cam", "Noodle pull stretched to camera, slow-mo", 1, "Iconic Asian food shot"),
    ("Wok Cam", "Wok being cleaned and re-seasoned (process video)", 2, "Educational + satisfying"),
    ("Chef Unhinged", "Chef rates customer orders out of 10 (with sass)", 2, "Personality builds following"),
    ("Chef Unhinged", "Chef tries a viral food hack and judges it", 2, "Reaction content = views"),
    ("Chef Unhinged", "Chef's least favourite food trend rant", 1, "Hot takes go viral"),
    ("Chef Unhinged", "Chef explains why one ingredient changes everything", 1, "Educational + opinion"),
    ("Chef Unhinged", "Chef ranks knives by personality", 2, "Random = memorable"),
    ("Chef Unhinged", "Chef cooks blindfolded for 60 seconds", 3, "Comedy + skill flex"),
    ("Chef Unhinged", "Chef's 'never order this in a restaurant' confession", 1, "Insider tea = saves"),
    ("Chef Unhinged", "Chef's morning prep routine in 30 sec", 2, "BTS = trust building"),
    ("Chef Unhinged", "Chef tries to recreate a customer's weird order", 2, "Wholesome chaos"),
    ("Chef Unhinged", "Chef's hot take: 'this dish is overrated'", 1, "Controversy = shares"),
    ("Customer Chaos", "First-bite reactions montage (5 customers)", 2, "Pure relatability"),
    ("Customer Chaos", "Customers describe Wokin in 3 words", 1, "Quick to shoot, high reach"),
    ("Customer Chaos", "'How spicy can you handle?' street interview style", 2, "Engaging format"),
    ("Customer Chaos", "Customers vote: best dish on the menu", 2, "UGC + social proof"),
    ("Customer Chaos", "Kid tries Pan Asian for the first time", 2, "Wholesome viral gold"),
    ("Customer Chaos", "Couple orders for each other (and judges)", 2, "Date-night relatable"),
    ("Customer Chaos", "Group challenge: who can finish first?", 2, "Competitive = engaging"),
    ("Customer Chaos", "Customer designs their own dish, chef makes it", 3, "Interactive content"),
    ("Customer Chaos", "Stranger gives us 5 stars or 1 star — gambling", 2, "Stakes = views"),
    ("Customer Chaos", "Regular customer hall of fame intro", 2, "Cult building"),
    ("Street Smart", "Origin of one menu item (60-sec history)", 2, "Educational saves"),
    ("Street Smart", "What 'wok hei' really means", 1, "Niche knowledge = authority"),
    ("Street Smart", "Why we use sesame oil at the END not the start", 1, "Practical tip = saves"),
    ("Street Smart", "Difference between 5 types of soy sauce", 2, "Educational + visual"),
    ("Street Smart", "Why authentic chow mein is different from yours", 1, "Myth-busting = controversy"),
    ("Street Smart", "How chopsticks were invented (animated)", 2, "Did-you-know format"),
    ("Street Smart", "What MSG actually does (and why it's fine)", 1, "Myth-busting saves"),
    ("Street Smart", "How to spot a fake 'authentic' Chinese restaurant", 1, "Sass + education"),
    ("Street Smart", "Why ginger is in everything (the science)", 1, "Health angle"),
    ("Street Smart", "The truth about fortune cookies", 2, "Fun fact = shares"),
    ("Meme Lord", "Trending audio + Wokin angle (whatever's hot)", 1, "Algorithm fuel"),
    ("Meme Lord", "Wokin dishes as Bollywood villains", 2, "Local relatable humour"),
    ("Meme Lord", "POV: When the chef adds 'extra spicy' anyway", 1, "Relatable trope"),
    ("Meme Lord", "Wokin staff vs customers chaos meme", 2, "Workplace humour"),
    ("Meme Lord", "'Things I tell myself before ordering 5 dishes'", 1, "Self-aware humour"),
    ("Meme Lord", "Wokin dishes as Karachi neighbourhoods", 2, "Local meme = local virality"),
    ("Meme Lord", "When you finish the noodles but there's still sauce left", 1, "Universal foodie pain"),
    ("Meme Lord", "Wokin order vs your bank account drama", 1, "Money meme = relatable"),
    ("Meme Lord", "Telling your friend 'one more bite' for the 7th time", 1, "Universal lie"),
    ("Meme Lord", "'My therapist vs my Wokin order' format", 1, "Mental health humour"),
    ("Wok Cam", "Egg drop into hot soup, super slow-mo", 1, "Hypnotic single shot"),
    ("Wok Cam", "Dumplings being folded by hand, time-lapse", 2, "Skill display"),
    ("Customer Chaos", "Couples rate each other's orders out of 10", 2, "Date format = shares"),
    ("Chef Unhinged", "Chef teaches customer to use chopsticks (cringe)", 2, "Wholesome content"),
    ("Street Smart", "Why noodle length means longevity (Chinese tradition)", 1, "Cultural depth"),
    ("Meme Lord", "When the spice hits 3 minutes later", 1, "Universal experience"),
    ("Wok Cam", "Crispy chilli oil being made — drizzle into bowl", 2, "Sensory satisfaction"),
    ("Customer Chaos", "Kid tries to pronounce dish names (wholesome)", 1, "Cuteness = saves"),
    ("Chef Unhinged", "Chef judges instant noodles brands", 2, "Brand opinion = engagement"),
    ("Street Smart", "Why we never serve fried rice with fresh rice", 1, "Pro tip = authority"),
    ("Meme Lord", "Friend group ordering chaos compilation", 1, "Universal friend trope"),
    ("Wok Cam", "Sauce reduction simmering loop", 1, "Aesthetic loop"),
]

for idx, (pillar, idea, effort, why) in enumerate(backup_reels):
    row = 4 + idx
    fill = CREAM if idx % 2 == 0 else WHITE
    pillar_colors = {
        "Wok Cam": "FFE5E5", "Chef Unhinged": "FFF0D9", "Customer Chaos": "E5F4FF",
        "Street Smart": "E8F5E9", "Meme Lord": "F3E5F5",
    }
    pf = pillar_colors.get(pillar, fill)
    style_cell(ws.cell(row=row, column=1, value=idx+1), fill=fill, bold=True, align="center")
    style_cell(ws.cell(row=row, column=2, value=pillar), fill=pf, bold=True, size=9, align="center")
    style_cell(ws.cell(row=row, column=3, value=idea), fill=fill, size=9)
    style_cell(ws.cell(row=row, column=4, value="★" * effort + "☆" * (5-effort)), fill=fill, size=10, align="center")
    style_cell(ws.cell(row=row, column=5, value=why), fill=fill, size=9)
    ws.row_dimensions[row].height = 32

ws.freeze_panes = "A4"
print("Backup Reels sheet done")

# ===== SHEET 5: BACKUP STORIES =====
ws = wb.create_sheet("Backup Stories")
ws.sheet_view.showGridLines = False

headers = ["#", "Type", "Story Idea", "Effort", "When To Use"]
widths = [6, 16, 60, 10, 30]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

title_row(ws, 1, "BACKUP STORY IDEAS — Quick & Lazy Friendly", len(headers), fill=BLACK)
title_row(ws, 2, "70+ low-effort story ideas. Mix and match daily.", len(headers), fill=RED, size=10)

for i, h in enumerate(headers, 1):
    style_header(ws.cell(row=3, column=i, value=h))
ws.row_dimensions[3].height = 28

backup_stories = [
    ("Poll", "Spicy or saucy?", 1, "Lunch hours"),
    ("Poll", "Noodles or rice?", 1, "Lunch / dinner"),
    ("Poll", "Half or full plate?", 1, "Pre-lunch"),
    ("Poll", "Veg or non-veg today?", 1, "Morning"),
    ("This or That", "Chow mein vs Hakka noodles", 1, "Lunch"),
    ("This or That", "Dim sum vs spring rolls", 1, "Snack hours"),
    ("This or That", "Szechuan vs Cantonese", 1, "Anytime"),
    ("This or That", "Crispy vs saucy", 1, "Anytime"),
    ("Quiz", "Guess the dish from the emoji", 1, "Engagement booster"),
    ("Quiz", "Spice tolerance test (1-10 scale)", 1, "Heat Test promo"),
    ("Quiz", "Which Wokin dish are you?", 2, "Personality bait"),
    ("Quiz", "Match the sauce to the dish", 1, "Educational"),
    ("Question", "What should we cook tomorrow?", 1, "Evening"),
    ("Question", "What's your unhinged food combo?", 1, "Anytime"),
    ("Question", "Drop your favourite Wokin moment", 1, "End of week"),
    ("Question", "What dish brings back memories?", 1, "Sunday vibes"),
    ("BTS", "Chef opening kitchen", 1, "Morning"),
    ("BTS", "Deliveries arriving fresh", 1, "Morning"),
    ("BTS", "Quick wok flame clip", 1, "Anytime"),
    ("BTS", "Plating in slow motion", 1, "Anytime"),
    ("BTS", "Closing time, last clean of the wok", 1, "Night"),
    ("BTS", "Staff prepping garnishes", 1, "Afternoon"),
    ("BTS", "Mid-rush chaos clip", 1, "Lunch / dinner"),
    ("BTS", "Quiet moment before service", 1, "Pre-rush"),
    ("UGC Reshare", "Reshare tagged photo with 'thanks for the love'", 1, "Daily habit"),
    ("UGC Reshare", "Reshare story tag with 'YESSS this dish'", 1, "Anytime"),
    ("UGC Reshare", "Compile UGC into a single slide", 2, "Weekly"),
    ("UGC Reshare", "Tag the customer back with reaction emoji", 1, "Daily"),
    ("Countdown", "To Heat Test sign-up close", 1, "Pre-event"),
    ("Countdown", "To weekend specials launch", 1, "Friday"),
    ("Countdown", "To new menu drop", 1, "Pre-launch"),
    ("Countdown", "To closing time on busy days", 1, "Evening"),
    ("Sneak Peek", "Tomorrow's reel — 2-second blurred clip", 1, "Daily"),
    ("Sneak Peek", "Test plating of an unreleased dish", 1, "Pre-launch"),
    ("Sneak Peek", "Chef testing a new sauce", 1, "Anytime"),
    ("Sneak Peek", "New ingredient just delivered", 1, "Morning"),
    ("Hot Take", "Text overlay: 'this is the best dish on the menu'", 1, "Anytime"),
    ("Hot Take", "'Stop ordering THIS, order THAT instead'", 1, "Lunch"),
    ("Hot Take", "'Soy sauce on rice is a war crime'", 1, "Anytime"),
    ("Stat", "Today we served X plates", 1, "End of day"),
    ("Stat", "Total chillies used this week", 1, "End of week"),
    ("Stat", "Most ordered dish today", 1, "End of day"),
    ("Stat", "Followers count milestone", 1, "Milestone day"),
    ("Throwback", "Old menu item — 'should this come back?'", 1, "Throwback Thursday"),
    ("Throwback", "First-day-of-business photo", 1, "Anniversary"),
    ("Sticker Game", "Slider: How spicy is your week?", 1, "Monday"),
    ("Sticker Game", "Emoji slider: How hungry are you?", 1, "Lunch"),
    ("Sticker Game", "Reaction slider: Rate this plate", 1, "Anytime"),
    ("Wokin Wrapped", "Top dish ordered this week", 1, "Sunday"),
    ("Wokin Wrapped", "Busiest hour stat", 1, "Sunday"),
    ("Wokin Wrapped", "Weirdest order received", 1, "Sunday"),
    ("Wokin Wrapped", "Customer of the week shoutout", 1, "Sunday"),
    ("Reel Promo", "Static frame from today's reel + arrow", 1, "Daily"),
    ("Reel Promo", "'New reel just dropped' graphic", 1, "Daily"),
    ("Sound Tease", "Just audio of a sizzle, no video", 1, "Anytime"),
    ("Sound Tease", "Chopsticks clinking ASMR", 1, "Anytime"),
    ("Sound Tease", "Knife on board rhythm", 1, "Anytime"),
    ("Visual Tease", "Steam rising, phone propped on counter", 1, "Anytime"),
    ("Visual Tease", "Sauce bottle silhouette in window light", 1, "Morning"),
    ("Visual Tease", "Empty plate with chopsticks placed neatly", 1, "End of meal"),
    ("Hype", "Heat Test contestant photo + countdown", 1, "Pre-test"),
    ("Hype", "'5 mins to closing — last orders' urgency", 1, "Closing time"),
    ("Joke", "'Me explaining the menu to my parents'", 1, "Anytime"),
    ("Joke", "'When the chilli kicks in 5 minutes later'", 1, "Anytime"),
    ("Educational", "Did you know: this ingredient comes from...", 1, "Anytime"),
    ("Educational", "Quick fact: how to hold chopsticks", 1, "Beginner-friendly"),
    ("Mood Check", "How's your Monday treating you? Slider", 1, "Monday"),
    ("Mood Check", "Friday vibes — emoji response", 1, "Friday"),
    ("Boomerang", "Sauce drizzle boomerang", 1, "Anytime"),
    ("Boomerang", "Chopsticks lifting noodles boomerang", 1, "Anytime"),
    ("Boomerang", "Cheers with drinks boomerang", 1, "Evening"),
]

type_colors = {
    "Poll": "E5F4FF", "This or That": "E5F4FF", "Quiz": "F3E5F5",
    "Question": "F3E5F5", "BTS": "FFE5E5", "UGC Reshare": "E8F5E9",
    "Countdown": "FFF0D9", "Sneak Peek": "FFF0D9", "Hot Take": "FFD6D6",
    "Stat": "E8F5E9", "Throwback": "FFF8DC", "Sticker Game": "F3E5F5",
    "Wokin Wrapped": "E8F5E9", "Reel Promo": "FFE5E5", "Sound Tease": "FFE5E5",
    "Visual Tease": "FFE5E5", "Hype": "FFD6D6", "Joke": "F3E5F5",
    "Educational": "E8F5E9", "Mood Check": "F3E5F5", "Boomerang": "FFE5E5",
}

for idx, (typ, idea, effort, when) in enumerate(backup_stories):
    row = 4 + idx
    fill = CREAM if idx % 2 == 0 else WHITE
    tf = type_colors.get(typ, fill)
    style_cell(ws.cell(row=row, column=1, value=idx+1), fill=fill, bold=True, align="center")
    style_cell(ws.cell(row=row, column=2, value=typ), fill=tf, bold=True, size=9, align="center")
    style_cell(ws.cell(row=row, column=3, value=idea), fill=fill, size=9)
    style_cell(ws.cell(row=row, column=4, value="★" * effort + "☆" * (5-effort)), fill=fill, size=9, align="center")
    style_cell(ws.cell(row=row, column=5, value=when), fill=fill, size=9)
    ws.row_dimensions[row].height = 24

ws.freeze_panes = "A4"
print("Backup Stories sheet done")

# ===== SHEET 6: HOOKS LIBRARY =====
ws = wb.create_sheet("Hooks Library")
ws.sheet_view.showGridLines = False
ws.column_dimensions['A'].width = 6
ws.column_dimensions['B'].width = 22
ws.column_dimensions['C'].width = 70
ws.column_dimensions['D'].width = 35

title_row(ws, 1, "HOOKS LIBRARY — Stop the Scroll in 1.5 Seconds", 4, fill=BLACK)
title_row(ws, 2, "Plug-and-play hook formulas + caption templates", 4, fill=RED, size=10)

for i, h in enumerate(["#", "Category", "Hook / Template", "Best For"], 1):
    style_header(ws.cell(row=3, column=i, value=h))
ws.row_dimensions[3].height = 28

hooks = [
    ("Hook", "POV: you walked into Wokin and ___", "Wok Cam reels"),
    ("Hook", "Things I'd never tell my customers but ___", "Chef Unhinged"),
    ("Hook", "Tell me you eat at Wokin without telling me ___", "Meme Lord"),
    ("Hook", "I tried [trend] using a wok and ___", "Trend hijack"),
    ("Hook", "This is illegal in [X] country but we do it anyway", "Hot take"),
    ("Hook", "Stop scrolling. You need to know this about [dish]", "Educational"),
    ("Hook", "Chef said don't post this but ___", "BTS"),
    ("Hook", "POV: the moment you taste real wok hei", "Wok Cam"),
    ("Hook", "Things only Wokin regulars notice", "Cult building"),
    ("Hook", "If your order is [X], we have to talk", "Personality bait"),
    ("Hook", "Watch what happens when oil hits the wok ___", "Wok Cam"),
    ("Hook", "We tested if [myth] is real and ___", "Street Smart"),
    ("Hook", "The dish nobody orders but should", "Underrated reveal"),
    ("Hook", "Customer said this and we had to film it", "Customer Chaos"),
    ("Hook", "Don't try this at home (we mean it)", "Skill flex"),
    ("Caption", "Your eardrums called. They want this on repeat. 🔊🔥", "ASMR reels"),
    ("Caption", "Tag the friend who orders the SAME dish every time.", "Tag bait"),
    ("Caption", "Bookmark this for your next date night order.", "Save bait"),
    ("Caption", "We said what we said. 🥢", "Hot take"),
    ("Caption", "If you know, you know. 🤝", "Insider"),
    ("Caption", "Drop a 🔥 if you survived this dish.", "Comment bait"),
    ("Caption", "POV: you finally tried the dish everyone's talking about.", "FOMO"),
    ("Caption", "Be honest — when did you last slurp like you meant it?", "Question hook"),
    ("Caption", "Send this to the friend who needs convincing.", "Share bait"),
    ("Caption", "Ranking our dishes worst to best. Don't unfollow us.", "Controversy"),
    ("Formula", "[Bold claim] + [emoji combo]", "Easy caption"),
    ("Formula", "[Question] + [tag the friend who...]", "Tag bait"),
    ("Formula", "[Visual] + [text overlay reveal at end]", "Reel structure"),
    ("Formula", "[Trending audio] + [Wokin context]", "Trend hijack"),
    ("Formula", "[Problem] + [Wokin = solution] + [emoji]", "Selling without selling"),
]

for idx, (cat, txt, best) in enumerate(hooks):
    row = 4 + idx
    fill = CREAM if idx % 2 == 0 else WHITE
    cf = {"Hook": "FFD6D6", "Caption": "E5F4FF", "Formula": "E8F5E9"}.get(cat, fill)
    style_cell(ws.cell(row=row, column=1, value=idx+1), fill=fill, bold=True, align="center")
    style_cell(ws.cell(row=row, column=2, value=cat), fill=cf, bold=True, size=9, align="center")
    style_cell(ws.cell(row=row, column=3, value=txt), fill=fill, size=10)
    style_cell(ws.cell(row=row, column=4, value=best), fill=fill, size=9)
    ws.row_dimensions[row].height = 28

ws.freeze_panes = "A4"
print("Hooks Library sheet done")

# ===== SHEET 7: HASHTAG BANK =====
ws = wb.create_sheet("Hashtag Bank")
ws.sheet_view.showGridLines = False
ws.column_dimensions['A'].width = 22
ws.column_dimensions['B'].width = 90

title_row(ws, 1, "HASHTAG BANK — Pre-Built Sets", 2, fill=BLACK)
title_row(ws, 2, "Use 8-12 max per post. Mix brand + niche + trending.", 2, fill=RED, size=10)

for i, h in enumerate(["Set", "Hashtags"], 1):
    style_header(ws.cell(row=3, column=i, value=h))
ws.row_dimensions[3].height = 28

hashtag_sets = [
    ("Brand Core", "#WokinKarachi #BoldFlavours #WokinFamily #WokToTable"),
    ("Wok Cam (ASMR)", "#FoodASMR #SatisfyingFood #FoodPorn #WokTok #FoodieGram #PakistanFoodie #KarachiEats #WokinKarachi"),
    ("Chef Unhinged", "#ChefLife #BehindTheKitchen #ChefHotTake #FoodieHumour #PakistaniChef #WokinKarachi #BoldFlavours"),
    ("Customer Chaos", "#FoodieReaction #FirstBite #SpicyChallenge #FoodieGram #KarachiFoodie #WokinKarachi #BoldFlavours"),
    ("Street Smart (educational)", "#FoodFacts #DidYouKnow #PanAsianFood #FoodHistory #FoodieEducation #WokinKarachi #ChefTips"),
    ("Meme Lord", "#FoodMemes #FoodieHumour #RelatableFoodie #PakistaniFoodie #KarachiEats #WokinKarachi #BoldFlavours"),
    ("Heat Test Challenge", "#SpicyChallenge #HeatTest #FoodChallenge #SpicyFood #ChilliChallenge #WokinHeatTest #KarachiEats #WokinKarachi"),
    ("Sunday Lazy Brunch", "#SundayBrunch #BrunchVibes #PanAsianBrunch #KarachiBrunch #WeekendVibes #WokinKarachi #BoldFlavours"),
    ("UGC Compilation", "#WokinFamily #CustomerLove #FoodieGram #KarachiEats #PakistanFoodie #WokinKarachi #BoldFlavours"),
    ("New Menu Drop", "#NewMenu #MenuDrop #FoodieAlert #KarachiEats #PakistanFoodie #WokinKarachi #BoldFlavours #JustDropped"),
    ("Throwback Thursday", "#ThrowbackThursday #TBT #FoodieMemories #WokinClassics #KarachiEats #WokinKarachi #BoldFlavours"),
]

for idx, (name, tags) in enumerate(hashtag_sets):
    row = 4 + idx
    fill = CREAM if idx % 2 == 0 else WHITE
    style_cell(ws.cell(row=row, column=1, value=name), fill=BLACK, bold=True, color=WHITE, size=10, align="center")
    style_cell(ws.cell(row=row, column=2, value=tags), fill=fill, size=10)
    ws.row_dimensions[row].height = 38

print("Hashtag Bank sheet done")

# ===== SHEET 8: TRACKER =====
ws = wb.create_sheet("Tracker")
ws.sheet_view.showGridLines = False
headers_t = ["Day", "Date", "Reel Title", "Posted?", "Views", "Likes", "Saves", "Shares", "Comments", "Save Rate %", "Notes"]
widths_t = [6, 12, 35, 9, 10, 10, 10, 10, 11, 12, 30]
for i, w in enumerate(widths_t, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

title_row(ws, 1, "PERFORMANCE TRACKER", len(headers_t), fill=BLACK)
title_row(ws, 2, "Fill in after posting. Save Rate auto-calculates.", len(headers_t), fill=RED, size=10)

for i, h in enumerate(headers_t, 1):
    style_header(ws.cell(row=3, column=i, value=h))
ws.row_dimensions[3].height = 28

for i in range(30):
    row = 4 + i
    fill = CREAM if i % 2 == 0 else WHITE
    style_cell(ws.cell(row=row, column=1, value=i+1), fill=fill, bold=True, align="center")
    date = datetime.date(2026, 4, 13) + datetime.timedelta(days=i)
    style_cell(ws.cell(row=row, column=2, value=date.strftime("%b %d")), fill=fill, align="center", size=9)
    for col in range(3, 12):
        style_cell(ws.cell(row=row, column=col, value=""), fill=fill, size=9, align="center")
    # Save rate formula
    ws.cell(row=row, column=10).value = f'=IF(E{row}>0,G{row}/E{row}*100,"")'
    ws.cell(row=row, column=10).number_format = '0.00"%"'
    ws.row_dimensions[row].height = 24

ws.freeze_panes = "C4"
print("Tracker sheet done")

wb.save("/home/user/wokin/Wokin_Social_Calendar.xlsx")
print("\nWORKBOOK SAVED")
